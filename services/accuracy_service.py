from __future__ import annotations

import re
from copy import copy
from dataclasses import dataclass
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from services.excel_processor import LedgerData, LedgerProcessingError


CHECK_ERROR_RE = re.compile(r"查\s*(\d+)\s*错\s*(\d+)")
FULLWIDTH_DIGITS = str.maketrans("０１２３４５６７８９", "0123456789")
APPENDED_ROW_FONT = Font(name="微软雅黑", sz=11, bold=False)
DATE_NUMBER_FORMAT = 'm"月"d"日"'
RATE_NUMBER_FORMAT = "0.00%"


@dataclass
class AccuracyUpdate:
    names: list[str]
    error_count: int | str
    street_name: str
    community_name: str
    location_name: str


class AccuracyService:
    def update_statistics(self, problem: LedgerData, statistics_path: str | Path) -> Path:
        workbook, export_name = self.build_updated_workbook(problem, statistics_path)
        source = Path(statistics_path)
        target = source.with_name(export_name)
        workbook.save(target)
        return target

    def build_updated_workbook(self, problem: LedgerData, statistics_path: str | Path) -> tuple[Workbook, str]:
        source = Path(statistics_path)
        if source.suffix.lower() != ".xlsx":
            raise LedgerProcessingError("准确率统计表仅支持 .xlsx 文件")

        date_label = problem.date_label
        if not date_label:
            raise LedgerProcessingError("无法从问题台账文件名识别日期")
        date_value = self._resolve_date_value(problem.metadata.get("date_value"), date_label)

        updates = self._collect_updates(problem)
        workbook = load_workbook(source)

        for sheet_name, records in updates.items():
            if sheet_name not in workbook.sheetnames:
                raise LedgerProcessingError(f"统计表缺少 Sheet：{sheet_name}")
            sheet = workbook[sheet_name]
            self._delete_empty_data_rows(sheet)
            name_col = self._find_name_col(sheet, sheet_name)
            error_col = self._find_accuracy_error_col(sheet)
            total_col = self._find_accuracy_total_col(sheet, error_col)
            rate_col = self._find_accuracy_rate_col(sheet, error_col)
            date_col = self._find_date_col(sheet)
            street_col = self._street_col(sheet, sheet_name)
            community_col = self._optional_header_col(sheet, "社区/村名称") if sheet_name == "小区" else None
            row_map = self._build_name_row_map(sheet, name_col)
            matched_updates: dict[int, int | str] = {}
            for record in records:
                row_index = self._match_location_row(row_map, record.names)
                if not row_index:
                    row_index = self._append_location_row(
                        sheet,
                        sheet_name,
                        record,
                        name_col,
                        total_col,
                        error_col,
                        rate_col,
                        date_col,
                        date_value,
                        date_label,
                    )
                    row_map[self._normalize_name(record.location_name)] = row_index
                    continue
                if row_index not in matched_updates:
                    matched_updates[row_index] = record.error_count
                elif isinstance(record.error_count, int):
                    current = matched_updates[row_index]
                    matched_updates[row_index] = (current if isinstance(current, int) else 0) + record.error_count
            for row_index, error_count in matched_updates.items():
                sheet.cell(row=row_index, column=error_col).value = error_count
                date_cell = sheet.cell(row=row_index, column=date_col)
                date_cell.value = date_value if date_value else date_label
                if date_value:
                    date_cell.number_format = DATE_NUMBER_FORMAT
                self._center_appended_location_cells(sheet, row_index, sheet_name, street_col, community_col, name_col)

        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True

        return workbook, f"{date_label}小区村值守率及投放准确率统计.xlsx"

    def _collect_updates(self, problem: LedgerData) -> dict[str, list[AccuracyUpdate]]:
        indexes = {name: self._header_index(problem.headers, name) for name in ("2级点位", "具体问题")}
        second_indicator_idx = self._optional_header_index(problem.headers, "2级指标")
        third_indicator_idx = self._optional_header_index(problem.headers, "3级指标")
        candidate_name_indexes = [
            problem.headers.index(name)
            for name in ("3级点位", "5级点位", "4级点位")
            if name in problem.headers
        ]
        if not candidate_name_indexes:
            raise LedgerProcessingError("缺少必要字段：3级点位")

        grouped: dict[str, dict[tuple[str, ...], dict[str, object]]] = {"小区": {}, "村居": {}}
        for row in problem.rows:
            point_type = self._cell_text(row, indexes["2级点位"])
            if point_type not in grouped:
                continue
            street_name = self._point_field_value(problem, row, "3级点位")
            community_name = self._point_field_value(problem, row, "4级点位")
            fifth_point = self._point_field_value(problem, row, "5级点位")
            location_name = fifth_point or community_name or street_name
            names = self._candidate_location_names(row, candidate_name_indexes)
            if not names:
                continue
            key = tuple(self._normalize_name(name) for name in names)
            record = grouped[point_type].setdefault(
                key,
                {
                    "names": names,
                    "check_error_sum": 0,
                    "has_check_error": False,
                    "has_fallback_match": False,
                    "fallback_count": 0,
                    "street_name": street_name,
                    "community_name": community_name,
                    "location_name": location_name,
                },
            )
            problem_text = self._cell_text(row, indexes["具体问题"])
            error_count = self._parse_error_count(problem_text)
            if error_count is not None:
                record["has_check_error"] = True
                record["check_error_sum"] = int(record["check_error_sum"]) + error_count

            second_indicator = self._cell_text(row, second_indicator_idx) if second_indicator_idx is not None else ""
            third_indicator = self._cell_text(row, third_indicator_idx) if third_indicator_idx is not None else ""
            is_resident_error = second_indicator == "居民自主投放" and third_indicator == "投放错误"
            is_station_problem = second_indicator == "垃圾分类驿站" and third_indicator != "无问题"
            if is_resident_error or is_station_problem:
                record["has_fallback_match"] = True
                record["fallback_count"] = int(record["fallback_count"]) + 1

        updates: dict[str, list[AccuracyUpdate]] = {"小区": [], "村居": []}
        for point_type, records in grouped.items():
            for record in records.values():
                error_value: int | str
                if record["has_check_error"]:
                    error_value = int(record["check_error_sum"])
                elif record["has_fallback_match"]:
                    error_value = int(record["fallback_count"])
                else:
                    error_value = "-"
                updates[point_type].append(
                    AccuracyUpdate(
                        list(record["names"]),
                        error_value,
                        str(record["street_name"]),
                        str(record["community_name"]),
                        str(record["location_name"]),
                    )
                )
        return updates

    def _parse_error_count(self, text: str) -> int | None:
        normalized = (
            text.translate(FULLWIDTH_DIGITS)
            .replace("（", "(")
            .replace("）", ")")
            .replace("，", ",")
            .replace("：", ":")
        )
        match = CHECK_ERROR_RE.search(normalized)
        if not match:
            return None
        return int(match.group(2))

    def _candidate_location_names(self, row: list[object], indexes: list[int]) -> list[str]:
        result: list[str] = []
        for idx in indexes:
            value = self._cell_text(row, idx)
            if value and value not in {"小区", "村居", "点位"} and value not in result:
                result.append(value)
        return result

    def _point_field_value(self, problem: LedgerData, row: list[object], field_name: str) -> str:
        index = self._optional_header_index(problem.headers, field_name)
        return self._cell_text(row, index) if index is not None else ""

    def _match_location_row(self, row_map: dict[str, int], names: list[str]) -> int | None:
        for name in names:
            row_index = row_map.get(self._normalize_name(name))
            if row_index:
                return row_index
        return None

    def _find_name_col(self, sheet: Worksheet, sheet_name: str) -> int:
        candidates = ["小区名称"] if sheet_name == "小区" else ["村居名称", "村名称"]
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 5)):
            for cell in row:
                if str(cell.value or "").strip() in candidates:
                    return cell.column
        raise LedgerProcessingError(f"{sheet_name} Sheet 未找到名称列")

    def _find_accuracy_error_col(self, sheet: Worksheet) -> int:
        header_row = self._find_header_row(sheet, "投放错误数")
        for cell in sheet[header_row]:
            if str(cell.value or "").strip() == "投放错误数":
                return cell.column
        raise LedgerProcessingError(f"{sheet.title} Sheet 未找到投放错误数列")

    def _find_accuracy_total_col(self, sheet: Worksheet, error_col: int) -> int:
        header_row = self._find_header_row(sheet, "投放总数")
        for col_index in range(error_col - 1, 0, -1):
            if str(sheet.cell(row=header_row, column=col_index).value or "").strip() == "投放总数":
                return col_index
        raise LedgerProcessingError(f"{sheet.title} Sheet 未找到投放总数列")

    def _find_accuracy_rate_col(self, sheet: Worksheet, error_col: int) -> int:
        header_row = self._find_header_row_at_col(sheet, "投放错误数", error_col)
        for col_index in range(error_col + 1, sheet.max_column + 1):
            value = str(sheet.cell(row=header_row, column=col_index).value or "").strip()
            if value == "投放准确率":
                return col_index
            if value == "投放总数":
                break
        raise LedgerProcessingError(f"{sheet.title} Sheet 未找到投放准确率列")

    def _find_header_row_at_col(self, sheet: Worksheet, text: str, column: int) -> int:
        for row_index in range(1, min(sheet.max_row, 5) + 1):
            if str(sheet.cell(row=row_index, column=column).value or "").strip() == text:
                return row_index
        raise LedgerProcessingError(f"{sheet.title} Sheet 未找到字段：{text}")

    def _find_date_col(self, sheet: Worksheet) -> int:
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 5)):
            for cell in row:
                if str(cell.value or "").strip() == "日期":
                    return cell.column
        raise LedgerProcessingError(f"{sheet.title} Sheet 未找到日期列")

    def _append_location_row(
        self,
        sheet: Worksheet,
        sheet_name: str,
        record: AccuracyUpdate,
        name_col: int,
        total_col: int,
        error_col: int,
        rate_col: int,
        date_col: int,
        date_value: date | None,
        date_label: str,
    ) -> int:
        row_index = sheet.max_row + 1
        template_row = self._last_data_row(sheet) or max(4, sheet.max_row)
        self._copy_row_style(sheet, template_row, row_index)
        self._apply_appended_row_font(sheet, row_index)

        serial_col = self._optional_header_col(sheet, "序号")
        street_col = self._street_col(sheet, sheet_name)
        community_col = self._optional_header_col(sheet, "社区/村名称") if sheet_name == "小区" else None

        if serial_col:
            sheet.cell(row=row_index, column=serial_col).value = self._next_serial_number(sheet, serial_col)
        sheet.cell(row=row_index, column=street_col).value = record.street_name
        if community_col:
            sheet.cell(row=row_index, column=community_col).value = record.community_name
        sheet.cell(row=row_index, column=name_col).value = record.location_name
        self._center_appended_location_cells(sheet, row_index, sheet_name, street_col, community_col, name_col)
        sheet.cell(row=row_index, column=total_col).value = 10
        sheet.cell(row=row_index, column=error_col).value = record.error_count
        self._write_accuracy_rate_formula(sheet, row_index, total_col, error_col, rate_col)
        date_cell = sheet.cell(row=row_index, column=date_col)
        resolved_date = self._resolve_date_value(date_value, date_label)
        date_cell.value = resolved_date if resolved_date else date_label
        if resolved_date:
            date_cell.number_format = DATE_NUMBER_FORMAT
        return row_index

    def _delete_empty_data_rows(self, sheet: Worksheet) -> None:
        for row_index in range(sheet.max_row, 3, -1):
            if self._is_empty_row(sheet, row_index):
                sheet.delete_rows(row_index)

    def _is_empty_row(self, sheet: Worksheet, row_index: int) -> bool:
        for col_index in range(1, sheet.max_column + 1):
            value = sheet.cell(row=row_index, column=col_index).value
            if value is not None and str(value).strip() != "":
                return False
        return True

    def _last_data_row(self, sheet: Worksheet) -> int | None:
        for row_index in range(sheet.max_row, 3, -1):
            if not self._is_empty_row(sheet, row_index):
                return row_index
        return None

    def _write_accuracy_rate_formula(
        self,
        sheet: Worksheet,
        row_index: int,
        total_col: int,
        error_col: int,
        rate_col: int,
    ) -> None:
        total_ref = f"{get_column_letter(total_col)}{row_index}"
        error_ref = f"{get_column_letter(error_col)}{row_index}"
        rate_cell = sheet.cell(row=row_index, column=rate_col)
        rate_cell.value = f'=IF(OR({total_ref}="",{total_ref}=0,{error_ref}="",{error_ref}="-"),"",1-{error_ref}/{total_ref})'
        rate_cell.number_format = RATE_NUMBER_FORMAT

    def _copy_row_style(self, sheet: Worksheet, source_row: int, target_row: int) -> None:
        if source_row >= target_row:
            return
        sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height
        sheet.row_dimensions[target_row].hidden = sheet.row_dimensions[source_row].hidden
        for col_index in range(1, sheet.max_column + 1):
            source_cell = sheet.cell(row=source_row, column=col_index)
            target_cell = sheet.cell(row=target_row, column=col_index)
            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format
            if source_cell.alignment:
                target_cell.alignment = copy(source_cell.alignment)
            if source_cell.border:
                target_cell.border = copy(source_cell.border)
            if source_cell.fill:
                target_cell.fill = copy(source_cell.fill)
            if source_cell.font:
                target_cell.font = copy(source_cell.font)
            if source_cell.protection:
                target_cell.protection = copy(source_cell.protection)

    def _apply_appended_row_font(self, sheet: Worksheet, row_index: int) -> None:
        for col_index in range(1, sheet.max_column + 1):
            sheet.cell(row=row_index, column=col_index).font = copy(APPENDED_ROW_FONT)

    def _center_appended_location_cells(
        self,
        sheet: Worksheet,
        row_index: int,
        sheet_name: str,
        street_col: int,
        community_col: int | None,
        name_col: int,
    ) -> None:
        if sheet_name != "小区":
            return
        for col_index in [street_col, community_col, name_col]:
            if not col_index:
                continue
            cell = sheet.cell(row=row_index, column=col_index)
            alignment = copy(cell.alignment)
            alignment.horizontal = "center"
            alignment.vertical = "center"
            cell.alignment = alignment

    def _resolve_date_value(self, value: object, date_label: str) -> date | None:
        if isinstance(value, date):
            return value
        if isinstance(value, str) and value.strip():
            parsed = self._parse_date_label(value.strip())
            if parsed:
                return parsed
        return self._parse_date_label(date_label)

    def _parse_date_label(self, text: str) -> date | None:
        if not text:
            return None
        patterns = [
            re.compile(r"(20\d{2})[-_.年]?(\d{1,2})[-_.月]?(\d{1,2})日?"),
            re.compile(r"(\d{1,2})月(\d{1,2})日?"),
        ]
        for pattern in patterns:
            match = pattern.search(text)
            if not match:
                continue
            if len(match.groups()) == 3:
                return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
            return date(date.today().year, int(match.group(1)), int(match.group(2)))
        return None

    def _next_serial_number(self, sheet: Worksheet, serial_col: int) -> int:
        values = [
            sheet.cell(row=row_index, column=serial_col).value
            for row_index in range(4, sheet.max_row + 1)
        ]
        numbers = [value for value in values if isinstance(value, int)]
        return (max(numbers) if numbers else 0) + 1

    def _street_col(self, sheet: Worksheet, sheet_name: str) -> int:
        candidates = ["街道/乡镇名称", "街乡镇\n名称", "街乡镇名称"]
        if sheet_name == "村居":
            candidates = ["街乡镇\n名称", "街乡镇名称", "街道/乡镇名称"]
        for candidate in candidates:
            col = self._optional_header_col(sheet, candidate)
            if col:
                return col
        raise LedgerProcessingError(f"{sheet.title} Sheet 未找到街乡镇列")

    def _optional_header_col(self, sheet: Worksheet, text: str) -> int | None:
        target = re.sub(r"\s+", "", text)
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 5)):
            for cell in row:
                value = re.sub(r"\s+", "", str(cell.value or ""))
                if value == target:
                    return cell.column
        return None

    def _find_header_row(self, sheet: Worksheet, text: str) -> int:
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 5)):
            for cell in row:
                if str(cell.value or "").strip() == text:
                    return cell.row
        raise LedgerProcessingError(f"{sheet.title} Sheet 未找到字段：{text}")

    def _build_name_row_map(self, sheet: Worksheet, name_col: int) -> dict[str, int]:
        result: dict[str, int] = {}
        for row_index in range(4, sheet.max_row + 1):
            name = self._normalize_name(sheet.cell(row=row_index, column=name_col).value)
            if name and name not in result:
                result[name] = row_index
        return result

    def _header_index(self, headers: list[str], name: str) -> int:
        try:
            return headers.index(name)
        except ValueError as exc:
            raise LedgerProcessingError(f"缺少必要字段：{name}") from exc

    def _optional_header_index(self, headers: list[str], name: str) -> int | None:
        try:
            return headers.index(name)
        except ValueError:
            return None

    def _cell_text(self, row: list[object], index: int) -> str:
        if index >= len(row) or row[index] is None:
            return ""
        return str(row[index]).strip()

    def _normalize_name(self, value: object) -> str:
        return re.sub(r"\s+", "", str(value or "")).strip()
