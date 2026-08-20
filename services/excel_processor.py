from __future__ import annotations

import re
from io import BytesIO
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from utils.paths import app_root


DELETE_FIELDS = [
    "定位",
    "整改照片",
    "责任单位",
    "检查员",
    "初审审核员",
    "复审审核员",
    "二维码",
    "整改天数",
]

POINT_FIELDS = ["2级点位", "3级点位"]
PHOTO_FIELD_RE = re.compile(r"^问题照片(\d+)$")
POINT_FIELDS_ALL = ["1级点位", "2级点位", "3级点位", "4级点位", "5级点位"]
INDICATOR_FIELDS = ["1级指标", "2级指标", "3级指标"]
IMAGE_WIDTH_PX = 74
IMAGE_HEIGHT_PX = 85
FIXED_COLUMN_WIDTHS = {
    "1级点位": 10.21,
    "2级点位": 10.21,
    "3级点位": 10.21,
    "4级点位": 10.21,
    "5级点位": 10.21,
    "1级指标": 10.21,
    "2级指标": 10.21,
    "3级指标": 10.21,
    "具体问题": 12.37,
    "上报时间": 12.37,
    "上报日期": 12.37,
    "报送时间": 12.37,
}
PHOTO_COLUMN_WIDTH = 10.21


@dataclass
class LedgerImage:
    row_index: int
    column_name: str
    data: bytes
    width: int | float | None = None
    height: int | float | None = None


@dataclass
class LedgerData:
    headers: list[str]
    rows: list[list[object]]
    source_path: Path | None = None
    date_label: str = ""
    metadata: dict[str, object] = field(default_factory=dict)
    images: list[LedgerImage] = field(default_factory=list)

    def preview_rows(self, limit: int = 100) -> list[list[object]]:
        return self.rows[:limit]


class LedgerProcessingError(RuntimeError):
    pass


class ExcelProcessor:
    def load_ledger(self, path: str | Path) -> LedgerData:
        file_path = Path(path)
        workbook_path = self._ensure_xlsx(file_path)
        workbook = load_workbook(workbook_path, data_only=False)
        sheet = workbook.active
        header_row, headers, metadata = self._read_headers(sheet)
        rows = self._read_rows(sheet, header_row, len(headers))
        images = self._read_images(sheet, headers, header_row)
        return LedgerData(
            headers=headers,
            rows=rows,
            source_path=file_path,
            date_label=self.extract_date_label(file_path.name),
            metadata={**metadata, "sheet_title": sheet.title, "date_value": self.extract_date(file_path.name)},
            images=images,
        )

    def make_base_ledger(self, data: LedgerData) -> LedgerData:
        self._require_headers(data.headers, ["具体问题", *POINT_FIELDS])
        kept_indexes = [
            idx for idx, header in enumerate(data.headers) if header not in DELETE_FIELDS
        ]
        headers = [data.headers[idx] for idx in kept_indexes]
        rows = [[row[idx] if idx < len(row) else None for idx in kept_indexes] for row in data.rows]
        self._renumber_rows(headers, rows)
        images = [
            image
            for image in data.images
            if image.column_name in headers and image.row_index < len(rows)
        ]

        headers, rows = self._move_indicator_after_points(headers, rows)
        headers, rows = self._move_report_time_after_problem(headers, rows)
        headers, rows = self._move_photos_to_tail(headers, rows)
        return LedgerData(headers, rows, data.source_path, data.date_label, data.metadata, images)

    def make_problem_ledger(self, data: LedgerData) -> LedgerData:
        problem_idx = self._header_index(data.headers, "具体问题")
        rows = []
        row_map: dict[int, int] = {}
        for old_index, row in enumerate(data.rows):
            if str(row[problem_idx] if problem_idx < len(row) else "").strip() == "无问题":
                continue
            row_map[old_index] = len(rows)
            rows.append(row)
        images = [
            LedgerImage(row_map[image.row_index], image.column_name, image.data, image.width, image.height)
            for image in data.images
            if image.row_index in row_map
        ]
        return LedgerData(data.headers[:], rows, data.source_path, data.date_label, data.metadata, images)

    def split_by_location(self, data: LedgerData) -> dict[str, LedgerData]:
        second_idx = self._header_index(data.headers, "2级点位")
        third_idx = self._header_index(data.headers, "3级点位")
        groups: dict[tuple[str, str], list[list[object]]] = {}
        row_maps: dict[tuple[str, str], dict[int, int]] = {}
        for old_index, row in enumerate(data.rows):
            second = self._safe_text(row, second_idx)
            third = self._safe_text(row, third_idx)
            if not second and not third:
                continue
            key = (second, third)
            groups.setdefault(key, []).append(row)
            row_maps.setdefault(key, {})[old_index] = len(groups[key]) - 1

        result: dict[str, LedgerData] = {}
        date_label = data.date_label or "未识别日期"
        third_types: dict[str, set[str]] = {}
        for second, third in groups:
            third_types.setdefault(third, set()).add(second)
        for (second, third), rows in groups.items():
            key = (second, third)
            row_map = row_maps[key]
            images = [
                LedgerImage(row_map[image.row_index], image.column_name, image.data, image.width, image.height)
                for image in data.images
                if image.row_index in row_map
            ]
            clean_third = self._safe_filename(third or "未命名点位")
            if second == "村居" and len(third_types.get(third, set())) > 1:
                filename = f"{date_label}{clean_third}村居.xlsx"
            elif second in {"村居", "小区"}:
                filename = f"{date_label}{clean_third}.xlsx"
            else:
                filename = f"{date_label}{self._safe_filename(second or '其他')}_{clean_third}.xlsx"
            split_rows = [row[:] for row in rows]
            self._renumber_rows(data.headers, split_rows)
            result[filename] = LedgerData(
                data.headers[:],
                split_rows,
                data.source_path,
                data.date_label,
                data.metadata,
                images,
            )
        return result

    def save_ledger(self, data: LedgerData, path: str | Path) -> Path:
        out_path = Path(path)
        if out_path.suffix.lower() != ".xlsx":
            out_path = out_path.with_suffix(".xlsx")
        out_path.parent.mkdir(parents=True, exist_ok=True)
        self._save_xlsx(data, out_path)
        return out_path

    def _save_xlsx(self, data: LedgerData, out_path: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "台账"
        self._write_sheet(sheet, data.headers, data.rows)
        self._format_sheet(sheet, data.headers, len(data.rows))
        self._merge_header_rows(sheet, data.headers)
        self._write_images(sheet, data)
        workbook.save(out_path)

    def save_split_ledgers(self, ledgers: dict[str, LedgerData], output_dir: str | Path) -> list[Path]:
        out_dir = Path(output_dir)
        out_dir.mkdir(parents=True, exist_ok=True)
        saved = []
        for filename, ledger in ledgers.items():
            saved.append(self.save_ledger(ledger, out_dir / filename))
        return saved

    def _ensure_xlsx(self, file_path: Path) -> Path:
        suffix = file_path.suffix.lower()
        if suffix == ".xlsx":
            return file_path
        raise LedgerProcessingError("仅支持 .xlsx 台账文件，请先自行转换后再导入")

    def _runtime_temp_root(self) -> Path:
        return app_root() / "runtime_tmp"

    def _read_headers(self, sheet: Worksheet) -> tuple[int, list[str], dict[str, object]]:
        buffered: list[tuple[int, list[str]]] = []
        for row_index in range(1, min(sheet.max_row, 10) + 1):
            values = self._header_row_values(sheet, row_index)
            if any(values):
                buffered.append((row_index, values))
            if self._looks_like_business_header(values):
                parent = buffered[-2][1] if len(buffered) >= 2 else []
                headers = self._canonical_headers(values, parent)
                return row_index, headers, {"multi_header": bool(parent)}
        for row_index, values in buffered:
            if any(values):
                return row_index, values, {"multi_header": False}
        raise LedgerProcessingError("未找到表头行")

    def _header_row_values(self, sheet: Worksheet, row_index: int) -> list[str]:
        values: list[str] = []
        for col_index in range(1, sheet.max_column + 1):
            value = sheet.cell(row=row_index, column=col_index).value
            if value is None:
                value = self._merged_cell_value(sheet, row_index, col_index)
            values.append(str(value).strip() if value is not None else "")
        return values

    def _merged_cell_value(self, sheet: Worksheet, row_index: int, col_index: int) -> object | None:
        for merged_range in sheet.merged_cells.ranges:
            if (
                merged_range.min_row <= row_index <= merged_range.max_row
                and merged_range.min_col <= col_index <= merged_range.max_col
            ):
                return sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
        return None

    def _looks_like_business_header(self, values: list[str]) -> bool:
        required = {"2级点位", "3级点位", "具体问题"}
        return required.issubset(set(values))

    def _canonical_headers(self, values: list[str], parent_values: list[str]) -> list[str]:
        headers: list[str] = []
        problem_photo_count = 0
        for idx, raw in enumerate(values):
            parent = parent_values[idx] if idx < len(parent_values) else ""
            if parent == "问题照片" and raw.startswith("图片"):
                problem_photo_count += 1
                headers.append(f"问题照片{problem_photo_count}")
            elif parent == "整改照片" and raw.startswith("图片"):
                headers.append("整改照片")
            elif parent == "二维码" and raw.startswith("图片"):
                headers.append("二维码")
            else:
                headers.append(raw)
        return headers

    def _read_rows(self, sheet: Worksheet, header_row: int, width: int) -> list[list[object]]:
        rows = []
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            values = list(row[:width])
            values.extend([None] * (width - len(values)))
            if any(value not in (None, "") for value in values):
                rows.append(values)
        return rows

    def _write_sheet(self, sheet: Worksheet, headers: list[str], rows: Iterable[list[object]]) -> None:
        row_list = list(rows)
        header_rows = self._display_header_rows(headers)
        for header_row in header_rows:
            sheet.append(header_row)
        for row in row_list:
            sheet.append(row)
        for idx, header in enumerate(headers, start=1):
            max_len = max([len(str(header)), *[len(str(r[idx - 1] or "")) for r in row_list]], default=10)
            width = self._column_width_for_header(header, max_len)
            sheet.column_dimensions[sheet.cell(row=1, column=idx).column_letter].width = width

    def _format_sheet(self, sheet: Worksheet, headers: list[str], data_row_count: int) -> None:
        header_count = len(self._display_header_rows(headers))
        normal_font = Font(name="宋体")
        header_font = Font(name="宋体", bold=True)
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_side = Side(style="thin", color="000000")
        content_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        header_fill = PatternFill(
            fill_type="solid",
            fgColor="BFBFBF",
        )
        max_row = header_count + data_row_count
        max_col = len(headers)
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = header_font if cell.row <= header_count else normal_font
                cell.alignment = alignment
        for row_index in range(1, max_row + 1):
            if row_index > header_count:
                sheet.row_dimensions[row_index].height = 63
            else:
                sheet.row_dimensions[row_index].height = 15
        for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.border = content_border
                if cell.row <= header_count:
                    cell.fill = header_fill

    def _merge_header_rows(self, sheet: Worksheet, headers: list[str]) -> None:
        header_rows = self._display_header_rows(headers)
        if len(header_rows) < 2:
            return
        max_col = len(headers)
        col = 1
        while col <= max_col:
            top_value = sheet.cell(row=1, column=col).value
            end_col = col
            while end_col + 1 <= max_col and sheet.cell(row=1, column=end_col + 1).value == top_value:
                end_col += 1
            if end_col > col:
                sheet.merge_cells(start_row=1, start_column=col, end_row=1, end_column=end_col)
            elif sheet.cell(row=2, column=col).value == top_value:
                sheet.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
            col = end_col + 1

    def _display_header_rows(self, headers: list[str]) -> list[list[str]]:
        has_structured = any(header in POINT_FIELDS_ALL or header in INDICATOR_FIELDS for header in headers)
        if not has_structured:
            return [headers, headers]
        top_row: list[str] = []
        second_row: list[str] = []
        for header in headers:
            top, second = self._display_header_cell(header)
            top_row.append(top)
            second_row.append(second)
        return [top_row, second_row]

    def _display_header_cell(self, header: str) -> tuple[str, str]:
        if header in POINT_FIELDS_ALL:
            return "点位", header
        if header in INDICATOR_FIELDS:
            return "指标", header
        photo_match = PHOTO_FIELD_RE.match(header)
        if photo_match:
            return "问题照片", f"图片{photo_match.group(1)}"
        if header == "整改照片":
            return "整改照片", "图片1"
        if header == "二维码":
            return "二维码", "图片1"
        return header, header

    def _read_images(self, sheet: Worksheet, headers: list[str], header_row: int) -> list[LedgerImage]:
        images = []
        for image in getattr(sheet, "_images", []):
            marker = getattr(getattr(image, "anchor", None), "_from", None)
            if marker is None:
                continue
            row_index = marker.row + 1 - header_row - 1
            col_index = marker.col
            if row_index < 0 or col_index < 0 or col_index >= len(headers):
                continue
            try:
                data = image._data()
            except Exception:
                continue
            images.append(
                LedgerImage(
                    row_index=row_index,
                    column_name=headers[col_index],
                    data=data,
                    width=getattr(image, "width", None),
                    height=getattr(image, "height", None),
                )
            )
        return images

    def _write_images(self, sheet: Worksheet, data: LedgerData) -> None:
        for ledger_image in data.images:
            if ledger_image.column_name not in data.headers:
                continue
            if ledger_image.row_index < 0 or ledger_image.row_index >= len(data.rows):
                continue
            image = OpenpyxlImage(BytesIO(ledger_image.data))
            image.width = IMAGE_WIDTH_PX
            image.height = IMAGE_HEIGHT_PX
            col = data.headers.index(ledger_image.column_name) + 1
            row = ledger_image.row_index + len(self._display_header_rows(data.headers)) + 1
            sheet.add_image(image, f"{get_column_letter(col)}{row}")

    def _require_headers(self, headers: list[str], required: list[str]) -> None:
        missing = [name for name in required if name not in headers]
        if missing:
            raise LedgerProcessingError("缺少必要字段：" + "、".join(missing))

    def _renumber_rows(self, headers: list[str], rows: list[list[object]]) -> None:
        number_names = ["编号", "序号"]
        target = next((name for name in number_names if name in headers), None)
        if not target:
            return
        number_idx = headers.index(target)
        for row_number, row in enumerate(rows, start=1):
            if number_idx < len(row):
                row[number_idx] = row_number

    def _header_index(self, headers: list[str], name: str) -> int:
        try:
            return headers.index(name)
        except ValueError as exc:
            raise LedgerProcessingError(f"缺少必要字段：{name}") from exc

    def _move_indicator_after_points(
        self, headers: list[str], rows: list[list[object]]
    ) -> tuple[list[str], list[list[object]]]:
        indicator_names = [h for h in headers if "指标" in h]
        if not indicator_names:
            return headers, rows
        point_after = next((name for name in reversed(POINT_FIELDS_ALL) if name in headers), "3级点位")
        return self._move_columns(headers, rows, indicator_names, after=point_after)

    def _move_report_time_after_problem(
        self, headers: list[str], rows: list[list[object]]
    ) -> tuple[list[str], list[list[object]]]:
        time_names = [h for h in ("上报时间", "上报日期", "报送时间") if h in headers]
        if not time_names:
            return headers, rows
        return self._move_columns(headers, rows, time_names[:1], after="具体问题")

    def _move_photos_to_tail(
        self, headers: list[str], rows: list[list[object]]
    ) -> tuple[list[str], list[list[object]]]:
        existing = [h for h in headers if PHOTO_FIELD_RE.match(h)]
        if not existing:
            return headers, rows
        return self._move_columns(headers, rows, existing, to_tail=True)

    def _move_columns(
        self,
        headers: list[str],
        rows: list[list[object]],
        moving: list[str],
        after: str | None = None,
        to_tail: bool = False,
    ) -> tuple[list[str], list[list[object]]]:
        source_indexes = [headers.index(name) for name in moving if name in headers]
        remaining_indexes = [idx for idx in range(len(headers)) if idx not in source_indexes]
        if to_tail:
            new_indexes = remaining_indexes + source_indexes
        else:
            if after not in headers:
                return headers, rows
            after_idx = headers.index(after)
            insert_at = remaining_indexes.index(after_idx) + 1
            new_indexes = remaining_indexes[:insert_at] + source_indexes + remaining_indexes[insert_at:]
        new_headers = [headers[idx] for idx in new_indexes]
        new_rows = [[row[idx] if idx < len(row) else None for idx in new_indexes] for row in rows]
        return new_headers, new_rows

    def _column_width_for_header(self, header: str, max_len: int) -> float:
        if PHOTO_FIELD_RE.match(header) or header in {"整改照片", "二维码"}:
            return PHOTO_COLUMN_WIDTH
        if header in FIXED_COLUMN_WIDTHS:
            return FIXED_COLUMN_WIDTHS[header]
        return min(max(max_len + 2, 10), 35)

    def _safe_text(self, row: list[object], index: int) -> str:
        if index >= len(row) or row[index] is None:
            return ""
        return str(row[index]).strip()

    def _safe_filename(self, value: str) -> str:
        return re.sub(r'[\\/:*?"<>|]+', "_", value).strip() or "未命名"

    def extract_date_label(self, filename: str) -> str:
        stem = Path(filename).stem
        parsed = self._extract_date_parts(stem)
        if parsed:
            _, month, day = parsed
            return f"{month}\u6708{day}\u65e5"
        return ""

    def extract_date(self, filename: str) -> date | None:
        stem = Path(filename).stem
        parsed = self._extract_date_parts(stem)
        if parsed:
            year, month, day = parsed
            return date(year, month, day)
        return None

    def _extract_date_parts(self, stem: str) -> tuple[int, int, int] | None:
        patterns = [
            re.compile("(20\\d{2})[-_.\u5e74]?(\\d{1,2})[-_.\u6708]?(\\d{1,2})\u65e5?"),
            re.compile("(\\d{1,2})\u6708(\\d{1,2})\u65e5?"),
        ]
        for pattern in patterns:
            match = pattern.search(stem)
            if not match:
                continue
            if len(match.groups()) == 3:
                return int(match.group(1)), int(match.group(2)), int(match.group(3))
            return date.today().year, int(match.group(1)), int(match.group(2))
        return None

    def save_split_ledgers(self, ledgers: dict[str, LedgerData], output_dir: str | Path) -> list[Path]:
        out_dir = Path(output_dir)
        out_dir.mkdir(parents=True, exist_ok=True)
        saved = []
        for filename, ledger in ledgers.items():
            saved.append(self.save_ledger(ledger, out_dir / filename))
        return saved

    def _ensure_xlsx(self, file_path: Path) -> Path:
        suffix = file_path.suffix.lower()
        if suffix == ".xlsx":
            return file_path
        raise LedgerProcessingError("仅支持 .xlsx 台账文件，请先自行转换后再导入")

    def _runtime_temp_root(self) -> Path:
        return app_root() / "runtime_tmp"

    def _read_headers(self, sheet: Worksheet) -> tuple[int, list[str], dict[str, object]]:
        buffered: list[tuple[int, list[str]]] = []
        for row_index in range(1, min(sheet.max_row, 10) + 1):
            values = self._header_row_values(sheet, row_index)
            if any(values):
                buffered.append((row_index, values))
            if self._looks_like_business_header(values):
                parent = buffered[-2][1] if len(buffered) >= 2 else []
                headers = self._canonical_headers(values, parent)
                return row_index, headers, {"multi_header": bool(parent)}
        for row_index, values in buffered:
            if any(values):
                return row_index, values, {"multi_header": False}
        raise LedgerProcessingError("未找到表头行")

    def _header_row_values(self, sheet: Worksheet, row_index: int) -> list[str]:
        values: list[str] = []
        for col_index in range(1, sheet.max_column + 1):
            value = sheet.cell(row=row_index, column=col_index).value
            if value is None:
                value = self._merged_cell_value(sheet, row_index, col_index)
            values.append(str(value).strip() if value is not None else "")
        return values

    def _merged_cell_value(self, sheet: Worksheet, row_index: int, col_index: int) -> object | None:
        for merged_range in sheet.merged_cells.ranges:
            if (
                merged_range.min_row <= row_index <= merged_range.max_row
                and merged_range.min_col <= col_index <= merged_range.max_col
            ):
                return sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
        return None

    def _looks_like_business_header(self, values: list[str]) -> bool:
        required = {"2级点位", "3级点位", "具体问题"}
        return required.issubset(set(values))

    def _canonical_headers(self, values: list[str], parent_values: list[str]) -> list[str]:
        headers: list[str] = []
        problem_photo_count = 0
        for idx, raw in enumerate(values):
            parent = parent_values[idx] if idx < len(parent_values) else ""
            if parent == "问题照片" and raw.startswith("图片"):
                problem_photo_count += 1
                headers.append(f"问题照片{problem_photo_count}")
            elif parent == "整改照片" and raw.startswith("图片"):
                headers.append("整改照片")
            elif parent == "二维码" and raw.startswith("图片"):
                headers.append("二维码")
            else:
                headers.append(raw)
        return headers

    def _read_rows(self, sheet: Worksheet, header_row: int, width: int) -> list[list[object]]:
        rows = []
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            values = list(row[:width])
            values.extend([None] * (width - len(values)))
            if any(value not in (None, "") for value in values):
                rows.append(values)
        return rows

    def _write_sheet(self, sheet: Worksheet, headers: list[str], rows: Iterable[list[object]]) -> None:
        row_list = list(rows)
        header_rows = self._display_header_rows(headers)
        for header_row in header_rows:
            sheet.append(header_row)
        for row in row_list:
            sheet.append(row)
        for idx, header in enumerate(headers, start=1):
            max_len = max([len(str(header)), *[len(str(r[idx - 1] or "")) for r in row_list]], default=10)
            width = self._column_width_for_header(header, max_len)
            sheet.column_dimensions[sheet.cell(row=1, column=idx).column_letter].width = width

    def _format_sheet(self, sheet: Worksheet, headers: list[str], data_row_count: int) -> None:
        header_count = len(self._display_header_rows(headers))
        normal_font = Font(name="宋体")
        header_font = Font(name="宋体", bold=True)
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_side = Side(style="thin", color="000000")
        content_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        header_fill = PatternFill(
            fill_type="solid",
            fgColor="BFBFBF",
        )
        max_row = header_count + data_row_count
        max_col = len(headers)
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = header_font if cell.row <= header_count else normal_font
                cell.alignment = alignment
        for row_index in range(1, max_row + 1):
            if row_index > header_count:
                sheet.row_dimensions[row_index].height = 63
            else:
                sheet.row_dimensions[row_index].height = 15
        for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.border = content_border
                if cell.row <= header_count:
                    cell.fill = header_fill

    def _merge_header_rows(self, sheet: Worksheet, headers: list[str]) -> None:
        header_rows = self._display_header_rows(headers)
        if len(header_rows) < 2:
            return
        max_col = len(headers)
        col = 1
        while col <= max_col:
            top_value = sheet.cell(row=1, column=col).value
            end_col = col
            while end_col + 1 <= max_col and sheet.cell(row=1, column=end_col + 1).value == top_value:
                end_col += 1
            if end_col > col:
                sheet.merge_cells(start_row=1, start_column=col, end_row=1, end_column=end_col)
            elif sheet.cell(row=2, column=col).value == top_value:
                sheet.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
            col = end_col + 1

    def _display_header_rows(self, headers: list[str]) -> list[list[str]]:
        has_structured = any(header in POINT_FIELDS_ALL or header in INDICATOR_FIELDS for header in headers)
        if not has_structured:
            return [headers, headers]
        top_row: list[str] = []
        second_row: list[str] = []
        for header in headers:
            top, second = self._display_header_cell(header)
            top_row.append(top)
            second_row.append(second)
        return [top_row, second_row]

    def _display_header_cell(self, header: str) -> tuple[str, str]:
        if header in POINT_FIELDS_ALL:
            return "点位", header
        if header in INDICATOR_FIELDS:
            return "指标", header
        photo_match = PHOTO_FIELD_RE.match(header)
        if photo_match:
            return "问题照片", f"图片{photo_match.group(1)}"
        if header == "整改照片":
            return "整改照片", "图片1"
        if header == "二维码":
            return "二维码", "图片1"
        return header, header

    def _read_images(self, sheet: Worksheet, headers: list[str], header_row: int) -> list[LedgerImage]:
        images = []
        for image in getattr(sheet, "_images", []):
            marker = getattr(getattr(image, "anchor", None), "_from", None)
            if marker is None:
                continue
            row_index = marker.row + 1 - header_row - 1
            col_index = marker.col
            if row_index < 0 or col_index < 0 or col_index >= len(headers):
                continue
            try:
                data = image._data()
            except Exception:
                continue
            images.append(
                LedgerImage(
                    row_index=row_index,
                    column_name=headers[col_index],
                    data=data,
                    width=getattr(image, "width", None),
                    height=getattr(image, "height", None),
                )
            )
        return images

    def _write_images(self, sheet: Worksheet, data: LedgerData) -> None:
        for ledger_image in data.images:
            if ledger_image.column_name not in data.headers:
                continue
            if ledger_image.row_index < 0 or ledger_image.row_index >= len(data.rows):
                continue
            image = OpenpyxlImage(BytesIO(ledger_image.data))
            image.width = IMAGE_WIDTH_PX
            image.height = IMAGE_HEIGHT_PX
            col = data.headers.index(ledger_image.column_name) + 1
            row = ledger_image.row_index + len(self._display_header_rows(data.headers)) + 1
            sheet.add_image(image, f"{get_column_letter(col)}{row}")

    def _require_headers(self, headers: list[str], required: list[str]) -> None:
        missing = [name for name in required if name not in headers]
        if missing:
            raise LedgerProcessingError("缺少必要字段：" + "、".join(missing))

    def _renumber_rows(self, headers: list[str], rows: list[list[object]]) -> None:
        number_names = ["编号", "序号"]
        target = next((name for name in number_names if name in headers), None)
        if not target:
            return
        number_idx = headers.index(target)
        for row_number, row in enumerate(rows, start=1):
            if number_idx < len(row):
                row[number_idx] = row_number

    def _header_index(self, headers: list[str], name: str) -> int:
        try:
            return headers.index(name)
        except ValueError as exc:
            raise LedgerProcessingError(f"缺少必要字段：{name}") from exc

    def _move_indicator_after_points(
        self, headers: list[str], rows: list[list[object]]
    ) -> tuple[list[str], list[list[object]]]:
        indicator_names = [h for h in headers if "指标" in h]
        if not indicator_names:
            return headers, rows
        point_after = next((name for name in reversed(POINT_FIELDS_ALL) if name in headers), "3级点位")
        return self._move_columns(headers, rows, indicator_names, after=point_after)

    def _move_report_time_after_problem(
        self, headers: list[str], rows: list[list[object]]
    ) -> tuple[list[str], list[list[object]]]:
        time_names = [h for h in ("上报时间", "上报日期", "报送时间") if h in headers]
        if not time_names:
            return headers, rows
        return self._move_columns(headers, rows, time_names[:1], after="具体问题")

    def _move_photos_to_tail(
        self, headers: list[str], rows: list[list[object]]
    ) -> tuple[list[str], list[list[object]]]:
        existing = [h for h in headers if PHOTO_FIELD_RE.match(h)]
        if not existing:
            return headers, rows
        return self._move_columns(headers, rows, existing, to_tail=True)

    def _move_columns(
        self,
        headers: list[str],
        rows: list[list[object]],
        moving: list[str],
        after: str | None = None,
        to_tail: bool = False,
    ) -> tuple[list[str], list[list[object]]]:
        source_indexes = [headers.index(name) for name in moving if name in headers]
        remaining_indexes = [idx for idx in range(len(headers)) if idx not in source_indexes]
        if to_tail:
            new_indexes = remaining_indexes + source_indexes
        else:
            if after not in headers:
                return headers, rows
            after_idx = headers.index(after)
            insert_at = remaining_indexes.index(after_idx) + 1
            new_indexes = remaining_indexes[:insert_at] + source_indexes + remaining_indexes[insert_at:]
        new_headers = [headers[idx] for idx in new_indexes]
        new_rows = [[row[idx] if idx < len(row) else None for idx in new_indexes] for row in rows]
        return new_headers, new_rows

    def _safe_text(self, row: list[object], index: int) -> str:
        if index >= len(row) or row[index] is None:
            return ""
        return str(row[index]).strip()

    def _safe_filename(self, value: str) -> str:
        return re.sub(r'[\\/:*?"<>|]+', "_", value).strip() or "未命名"
