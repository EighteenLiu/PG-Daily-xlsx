from pathlib import Path
import base64
from datetime import datetime

from openpyxl.drawing.image import Image
from openpyxl import Workbook, load_workbook

from services.excel_processor import ExcelProcessor, LedgerData
from services.accuracy_service import AccuracyService
from services.notice_service import NoticeService
from services.resident_service import ResidentService


def make_sample(path: Path) -> None:
    headers = [
        "编号",
        "1级点位",
        "2级点位",
        "3级点位",
        "定位",
        "指标名称",
        "具体问题",
        "上报时间",
        "问题照片1",
        "问题照片2",
        "问题照片3",
        "问题照片4",
        "整改照片",
        "责任单位",
        "检查员",
        "初审审核员",
        "复审审核员",
        "二维码",
        "整改天数",
        "备注1",
        "备注2",
        "备注3",
        "备注4",
    ]
    rows = [
        [20260706000048, "A", "村居", "新华", "loc", "分类投放", "桶满冒", "2026-07-08", "p1", "p2", "p3", "p4", "z", "r", "c", "a", "b", "q", 2, 1, 2, 3, 4],
        [20260706000049, "A", "小区", "阳光小区", "loc", "环境", "无问题", "2026-07-08", "p1", "p2", "p3", "p4", "z", "r", "c", "a", "b", "q", 2, 1, 2, 3, 4],
        [20260706000050, "A", "小区", "花园", "loc", "环境", "垃圾落地", "2026-07-08", "p1", "p2", "p3", "p4", "z", "r", "c", "a", "b", "q", 2, 1, 2, 3, 4],
    ]
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    image_path = path.parent / "tiny.png"
    image_path.write_bytes(
        base64.b64decode(
            "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMCAO+/p9sAAAAASUVORK5CYII="
        )
    )
    sheet.add_image(Image(image_path), "I2")
    sheet.add_image(Image(image_path), "I4")
    workbook.save(path)


def test_pipeline(tmp_path: Path) -> None:
    source = tmp_path / "7月8日原始台账.xlsx"
    make_sample(source)

    processor = ExcelProcessor()
    original = processor.load_ledger(source)
    base = processor.make_base_ledger(original)
    problem = processor.make_problem_ledger(base)
    split = processor.split_by_location(problem)

    assert len(base.headers) == 15
    assert [row[base.headers.index("编号")] for row in base.rows] == [1, 2, 3]
    assert "定位" not in base.headers
    assert "整改照片" not in base.headers
    assert base.headers[base.headers.index("3级点位") + 1] == "指标名称"
    assert base.headers[base.headers.index("具体问题") + 1] == "上报时间"
    assert base.headers[-4:] == ["问题照片1", "问题照片2", "问题照片3", "问题照片4"]
    assert len(problem.rows) == 2
    assert [row[problem.headers.index("编号")] for row in problem.rows] == [1, 3]
    assert len(problem.images) == 2
    assert set(split) == {"7月8日新华.xlsx", "7月8日花园.xlsx"}

    out_dir = tmp_path / "out"
    saved = processor.save_split_ledgers(split, out_dir)
    assert len(saved) == 2
    exported = load_workbook(out_dir / "7月8日新华.xlsx")
    sheet = exported.active
    assert sheet.max_row == 3
    assert sheet["A1"].value == "编号"
    assert [sheet.cell(row=2, column=idx).value for idx in range(2, 4)] == ["1级点位", "2级点位"]
    assert sheet.row_dimensions[1].height == 15
    assert sheet.row_dimensions[2].height == 15
    assert sheet.row_dimensions[3].height == 63
    assert sheet.column_dimensions["F"].width == 12.37
    assert sheet.column_dimensions["G"].width == 12.37
    assert sheet["A1"].font.name == "宋体"
    assert sheet["A1"].font.bold is True
    assert sheet["A3"].alignment.horizontal == "center"
    assert sheet["A3"].alignment.vertical == "center"
    assert sheet["A3"].alignment.wrap_text is True
    assert "A1:A2" in {str(item) for item in sheet.merged_cells.ranges}
    assert "B1:D1" in {str(item) for item in sheet.merged_cells.ranges}
    assert sheet["A1"].border.left.style == "thin"
    assert sheet["O3"].border.right.style == "thin"
    assert sheet["L3"].border.top.style == "thin"
    assert sheet["L3"].border.bottom.style == "thin"
    assert sheet["A1"].fill.fgColor.rgb == "00BFBFBF"
    assert sheet["B2"].fill.fgColor.rgb == "00BFBFBF"
    assert sheet["B3"].fill.fill_type is None
    assert exported.active._images[0].anchor.ext.cx == 74 * 9525
    assert exported.active._images[0].anchor.ext.cy == 85 * 9525

    base_path = processor.save_ledger(base, tmp_path / "base.xlsx")
    base_exported = load_workbook(base_path).active
    assert base_exported["A1"].fill.fgColor.rgb == "00BFBFBF"
    assert base_exported["B2"].fill.fgColor.rgb == "00BFBFBF"
    assert base_exported["L1"].fill.fgColor.rgb == "00BFBFBF"
    assert base_exported["O2"].fill.fgColor.rgb == "00BFBFBF"
    assert base_exported["A3"].fill.fill_type is None
    assert base_exported.merged_cells.ranges

    structured_path = tmp_path / "structured.xlsx"
    structured = Workbook()
    structured_sheet = structured.active
    structured_sheet.append(["编号", "1级点位", "2级点位", "3级点位", "4级点位", "5级点位", "1级指标", "2级指标", "3级指标", "具体问题", "上报时间", "问题照片1", "问题照片2", "问题照片3", "问题照片4"])
    structured_sheet.append([1, "点位", "小区", "街道A", "社区A", "很长很长的五级点位名称", "垃圾分类", "居民自主投放", "很长很长的三级指标名称", "很长很长的具体问题描述", "2026-07-08", None, None, None, None])
    structured.save(structured_path)
    structured_export = tmp_path / "structured_out.xlsx"
    processor.save_ledger(processor.load_ledger(structured_path), structured_export)
    structured_ws = load_workbook(structured_export).active
    for column in ["B", "C", "D", "E", "F", "G", "H", "I", "L", "M", "N", "O"]:
        assert structured_ws.column_dimensions[column].width == 10.21
    assert structured_ws.column_dimensions["F"].width == 10.21
    assert structured_ws.column_dimensions["I"].width == 10.21
    assert structured_ws.column_dimensions["J"].width == 12.37


def test_resident_ledger_filters_only_second_indicator(tmp_path: Path) -> None:
    source = tmp_path / "20260708原始台账.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    headers = ["编号", "1级点位", "2级点位", "3级点位", "具体问题", "1级指标", "2级指标", "3级指标"]
    sheet.append(headers)
    sheet.append([1, "点位", "小区", "街道A", "桶满冒", "垃圾分类", "居民自主投放", "投放错误"])
    sheet.append([2, "点位", "小区", "街道A", "满冒", "垃圾分类", "容器满冒", "容器垃圾满冒堆积"])
    workbook.save(source)
    processor = ExcelProcessor()
    problem = processor.make_problem_ledger(processor.make_base_ledger(processor.load_ledger(source)))

    resident = ResidentService().make_resident_ledger(problem)

    indicator_idx = resident.headers.index("2级指标")
    assert len(resident.rows) == 1
    assert all(row[indicator_idx] == "居民自主投放" for row in resident.rows)


def test_accuracy_statistics_updates_by_name_only(tmp_path: Path) -> None:
    problem_path = tmp_path / "20260708问题台账.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["编号", "1级点位", "2级点位", "3级点位", "4级点位", "5级点位", "1级指标", "2级指标", "3级指标", "具体问题"])
    sheet.append([1, "点位", "小区", "街道A", "社区A", "阳光小区", "垃圾分类", "居民自主投放", "投放错误", "垃圾分类驿站桶错误（查5错1）"])
    sheet.append([2, "点位", "小区", "街道A", "社区A", "阳光小区", "垃圾分类", "居民自主投放", "投放错误", "垃圾分类驿站桶错误（查 １０ 错 ２）"])
    sheet.append([3, "点位", "小区", "街道A", "社区A", "阳光小区", "垃圾分类", "居民自主投放", "投放错误", "普通投放错误"])
    sheet.append([4, "点位", "小区", "街道A", "社区B", "花园小区", "垃圾分类", "居民自主投放", "投放错误", "普通投放错误"])
    sheet.append([5, "点位", "小区", "街道A", "社区B", "花园小区", "垃圾分类", "居民自主投放", "投放错误", "普通投放错误"])
    sheet.append([6, "点位", "小区", "街道A", "社区C", "无居民小区", "垃圾分类", "容器满冒", "容器垃圾满冒堆积", "垃圾桶满冒"])
    sheet.append([7, "点位", "小区", "街道A", "社区D", "驿站小区", "垃圾分类", "垃圾分类驿站", "桶错误", "垃圾分类驿站桶错误"])
    sheet.append([8, "点位", "小区", "街道A", "社区D", "驿站小区", "垃圾分类", "垃圾分类驿站", "无问题", "无问题"])
    sheet.append([9, "点位", "村居", "镇A", "村A", "幸福村", "垃圾分类", "居民自主投放", "投放错误", "垃圾分类驿站桶错误（查35错4）"])
    sheet.append([10, "点位", "村居", "镇A", "村B", "忽略村", "垃圾分类", "容器满冒", "容器垃圾满冒堆积", "普通问题"])
    workbook.save(problem_path)

    stats_path = tmp_path / "统计表.xlsx"
    stats = Workbook()
    xiaoqu = stats.active
    xiaoqu.title = "小区"
    xiaoqu.append(["192个小区垃圾桶站值守率及投放准确率"])
    xiaoqu.append(["基本信息", None, None, None, None, "值守率", None, None, "投放准确率", None, None, "垃圾分类纯净率", None, None, "日期"])
    xiaoqu.append(["序号", "街道/乡镇名称", "社区/村名称", "小区名称", "小区\n总户数", "检查桶站数量", "值守人员数量", "值守率", "投放总数", "投放错误数", "投放准确率", "投放总数", "投放错误数", "投放准确率", None])
    xiaoqu.append([1, "街道A", "社区A", "阳光小区", 100, 1, "-", None, 10, 0, None, "-", "-", None, None])
    xiaoqu.append([2, "街道A", "社区B", "花园小区", 100, 1, "-", None, 10, 0, None, "-", "-", None, None])
    xiaoqu.append([3, "街道A", "社区C", "无居民小区", 100, 1, "-", None, 10, 7, None, "-", "-", None, None])
    xiaoqu.append([4, "街道A", "社区D", "驿站小区", 100, 1, "-", None, 10, 0, None, "-", "-", None, None])
    cunju = stats.create_sheet("村居")
    cunju.append(["267个村居垃圾桶站值守率及投放准确率"])
    cunju.append(["基本信息", None, None, "投放准确率", None, None, "垃圾分类纯净率", None, None, "日期"])
    cunju.append(["序号", "街乡镇\n名称", "村居名称", "投放总数", "投放错误数", "投放准确率", "投放总数", "投放错误数", "投放准确率", None])
    cunju.append([1, "镇A", "幸福村", 10, 0, None, "-", "-", None, None])
    cunju.append([2, "镇A", "忽略村", 10, 6, None, "-", "-", None, None])
    stats.save(stats_path)

    processor = ExcelProcessor()
    problem = processor.load_ledger(problem_path)
    saved = AccuracyService().update_statistics(problem, stats_path)

    updated = load_workbook(saved, data_only=False)
    assert updated["小区"].cell(row=4, column=10).value == 3
    assert updated["小区"].cell(row=4, column=15).value == datetime(2026, 7, 8)
    assert updated["小区"].cell(row=4, column=15).number_format == "m月d日"
    assert updated["小区"].cell(row=5, column=10).value == 2
    assert updated["小区"].cell(row=5, column=15).value == datetime(2026, 7, 8)
    assert updated["小区"].cell(row=6, column=10).value == "-"
    assert updated["小区"].cell(row=6, column=15).value == datetime(2026, 7, 8)
    assert updated["小区"].cell(row=7, column=10).value == 1
    assert updated["小区"].cell(row=7, column=15).value == datetime(2026, 7, 8)
    assert updated["村居"].cell(row=4, column=5).value == 4
    assert updated["村居"].cell(row=4, column=10).value == datetime(2026, 7, 8)
    assert updated["村居"].cell(row=4, column=10).number_format == "m月d日"
    assert updated["村居"].cell(row=5, column=5).value == "-"
    assert updated["村居"].cell(row=5, column=10).value == datetime(2026, 7, 8)


def test_notice_generation_uses_two_streets_and_one_town(tmp_path: Path) -> None:
    problem = LedgerData(
        headers=["编号", "1级点位", "2级点位", "3级点位", "4级点位", "5级点位", "具体问题", "1级指标", "2级指标", "3级指标"],
        rows=[
            ["", "点位", "小区", "滨河街道", "社区A", "小区甲", "问题1", "垃圾分类", "容器满冒", "容器垃圾满冒堆积"],
            ["", "点位", "小区", "滨河街道", "社区B", "小区乙", "问题2", "垃圾分类", "容器满冒", "容器垃圾满冒堆积"],
            ["", "点位", "小区", "滨河街道", "社区C", "小区丙", "问题3", "垃圾分类", "桶站及周边环境", "桶站及周边环境不整洁"],
            ["", "点位", "小区", "兴谷街道", "社区C", "小区丙", "问题4", "垃圾分类", "居民自主投放", "投放错误"],
            ["", "点位", "小区", "平谷镇", "社区D", "小区丁", "问题5", "垃圾分类", "容器整洁率", "容器脏污"],
            ["", "点位", "小区", "平谷镇", "社区E", "小区戊", "问题6", "垃圾分类", "容器整洁率", "容器脏污"],
        ],
        date_label="7月10日",
    )

    saved = NoticeService().generate_notice(problem, tmp_path / "通告.txt")
    text = saved.read_text(encoding="utf-8")

    assert "7月10日，区级检查滨河街道3个小区，发现问题3处" in text
    assert "容器垃圾满冒堆积（66.7%）、桶站及周边环境不整洁（33.3%）" in text
    assert "小区甲、小区乙和小区丙问题占比均为33.3%" in text
    assert "检查兴谷街道1个小区，发现问题1处" in text
    assert "居民自主投放不规范（100.0%）。其中小区丙问题率最高，占比100.0%" in text
    assert "检查平谷镇2个小区，发现问题2处" in text
    assert "容器脏污（100.0%）。其中小区丁和小区戊问题占比均为50.0%" in text
    assert "无（0.0%）" not in text
